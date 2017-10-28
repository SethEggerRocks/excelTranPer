import openpyxl


glc_wb = openpyxl.load_workbook('/Users/nichodeturbo/Desktop/Grateful Life Master Roster 170831.xlsx')
glc_sheet = glc_wb.get_sheet_by_name("Residential")
glc_otherThing = glc_sheet.max_row


do_wb = openpyxl.load_workbook('/Users/nichodeturbo/Desktop/Dougherty House IOP Master Roster August 2017.xlsx')
do_sheet = do_wb.get_sheet_by_name("OUTPATIENT")
do_otherThing = do_sheet.max_row


dr_wb = openpyxl.load_workbook('/Users/nichodeturbo/Desktop/Master Droege Daily Roster July 2017.xlsx')
dr_sheet = dr_wb.get_sheet_by_name("Residential")
dr_otherThing = dr_sheet.max_row


de_wb = openpyxl.load_workbook('/Users/nichodeturbo/Desktop/Master Droege Detox Daily Roster july 2017.xlsx')
de_sheet = de_wb.get_sheet_by_name("DETOX")
de_otherThing = de_sheet.max_row


wb = openpyxl.load_workbook('/Users/nichodeturbo/Desktop/Wrap Daily Roster 20170801master.xlsx')
sheet = wb.get_sheet_by_name("Residential")
otherThing = sheet.max_row


glc_emptyList = []
glc_asaList = []
glc_completedList = []
glc_arrestedList = []
glc_dischargedList = []


do_emptyList = []
do_asaList = []
do_completedList = []
do_arrestedList = []
do_dischargedList = []


dr_emptyList = []
dr_asaList = []
dr_completedList = []
dr_arrestedList = []
dr_dischargedList = []


de_emptyList = []
de_asaList = []
de_completedList = []
de_arrestedList = []
de_dischargedList = []


emptyList = []
asaList = []
completedList = []
arrestedList = []
dischargedList = []


def glcNumbers():
    for columnOfCellObjects in glc_sheet["J134" : "J155"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "ASA":
                glc_emptyList.append(cellObj.value)
            elif cellObj.value == "Discharged":
                glc_emptyList.append(cellObj.value)
            elif cellObj.value == "Completed":
                glc_emptyList.append(cellObj.value)



    for i in glc_emptyList:
        if i == "ASA":
            glc_asaList.append(i)
    print(" ")
    print("ASA percentage for GLC:" + str(len(glc_asaList)/len(glc_emptyList)* 100) + str("%"))

    for i in glc_emptyList:
        if i == "Discharged":
            glc_dischargedList.append(i)
    print("Discharged percentage for GLC:" + str(len(glc_dischargedList)/len(glc_emptyList) * 100) + str("%"))

    for i in glc_emptyList:
        if i == "Completed":
            glc_completedList.append(i)
    print("Graduated percentage for GLC:" + str(len(glc_completedList)/len(glc_emptyList) * 100) + str("%"))

def doughertyNumbers():
    for columnOfCellObjects in do_sheet["I93" : "I98"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "GRAD":
                do_emptyList.append(cellObj.value)
            elif cellObj.value == "Unsucc":
                do_emptyList.append(cellObj.value)


    for i in do_emptyList:
        if i == "GRAD":
            do_completedList.append(i)
    print(" ")
    print("Graduated percentage for IOP:" + str(len(do_completedList)/len(do_emptyList) * 100) + str("%"))

    for i in do_emptyList:
        if i == "Unsucc":
            do_dischargedList.append(i)
    print("Unsuccessful percentage for IOP:" + str(len(do_dischargedList)/len(do_emptyList) * 100) + str("%"))

def droegeNumbers():
    for columnOfCellObjects in dr_sheet["I58" : "I83"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "ASA":
                dr_emptyList.append(cellObj.value)
            elif cellObj.value == "Graduate":
                dr_emptyList.append(cellObj.value)
            elif cellObj.value == "Complete":
                dr_emptyList.append(cellObj.value)
            elif cellObj.value == "Discharged":
                dr_emptyList.append(cellObj.value)

    for i in dr_emptyList:
        if i == "ASA":
            dr_asaList.append(i)
    print(" ")
    print("ASA percentage for Droege:" + str(len(dr_asaList)/len(dr_emptyList) * 100) + str("%"))

    for i in dr_emptyList:
        if i == "Graduate":
            dr_completedList.append(i)
    print("Graduated percentage for Droege:" + str(len(dr_completedList)/len(dr_emptyList) * 100) + str("%"))

    for i in dr_emptyList:
        if i == "Complete":
            dr_arrestedList.append(i)
    print("Completion percentage for Droege:" + str(len(dr_arrestedList)/len(dr_emptyList) * 100) + str("%"))

    for i in dr_emptyList:
        if i == "Discharged":
            dr_dischargedList.append(i)
    print("Discharged percentage for Droege:" + str(len(dr_dischargedList)/len(dr_emptyList) * 100) + str("%"))

def detoxNumbers():
    for columnOfCellObjects in de_sheet["I25": "I113"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "ASA":
                de_emptyList.append(cellObj.value)
            elif cellObj.value == "Med Discharge":
                de_emptyList.append(cellObj.value)
            elif cellObj.value == "comp":
                de_emptyList.append(cellObj.value)
            elif cellObj.value == "Discharged":
                de_emptyList.append(cellObj.value)

    for i in de_emptyList:
        if i == "ASA":
            de_asaList.append(i)
    print(" ")
    print("ASA percentage for Detox:" + str(len(de_asaList) / len(de_emptyList) * 100) + str("%"))

    for i in de_emptyList:
        if i == "Med Discharge":
            de_completedList.append(i)
    print("MEd Discharge percentage for Detox:" + str(len(de_completedList) / len(de_emptyList) * 100) + str("%"))

    for i in de_emptyList:
        if i == "comp":
            de_arrestedList.append(i)
    print("Completion percentage for Detox:" + str(len(de_arrestedList) / len(de_emptyList) * 100) + str("%"))

    for i in de_emptyList:
        if i == "Discharged":
            de_dischargedList.append(i)
    print("Discharged percentage for Detox:" + str(len(de_dischargedList) / len(de_emptyList) * 100) + str("%"))

def wrapNumbers():
    for columnOfCellObjects in sheet["H55": "H84"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "ASA":
                emptyList.append(cellObj.value)
            elif cellObj.value == "Completed":
                emptyList.append(cellObj.value)
            elif cellObj.value == "Arrested":
                emptyList.append(cellObj.value)
            elif cellObj.value == "Discharged":
                emptyList.append(cellObj.value)

    for i in emptyList:
        if i == "ASA":
            asaList.append(i)
    print(" ")
    print("ASA percentage for WRAP:" + str(len(asaList)/len(emptyList) * 100) + str("%"))

    for i in emptyList:
        if i == "Completed":
            completedList.append(i)
    print("Completed percentage for WRAP:" + str(len(completedList)/len(emptyList) * 100) + str("%"))

    for i in emptyList:
        if i == "Arrested":
            arrestedList.append(i)
    print("Arrested percentage for WRAP:" + str(len(arrestedList)/len(emptyList) * 100) + str("%"))

    for i in emptyList:
        if i == "Discharged":
            dischargedList.append(i)
    print("Discharged percentage for WRAP:" + str(len(dischargedList)/len(emptyList) * 100) + str("%"))

    for i in emptyList:
        if i == "Un Succ Dis":
            dischargedList.append(i)
    print("Un Successful Discharge percentage for WRAP:" + str(len(dischargedList)/len(emptyList) * 100) + str("%"))


